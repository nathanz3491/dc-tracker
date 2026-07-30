"""ISO interconnection queue ingest — candidate generation, not a project feed.

Named `pjm.py` per the PRD's file layout, but handles all four ISOs via
`iso_maps`. What this module can and cannot honestly produce is documented at
length in `iso_maps.py`; the short version is that these are **generator**
queues with no data-center column, so:

* A match is a keyword guess. ``IngestRecord.confidence_cap`` is 1.
* Queue MW is generator nameplate. It goes to ``notes``, not
  ``project.mw_planned``, unless the operator passes ``trust_gen_mw=True``.
* Location is county-granular, so it populates ``project.county`` and cannot
  auto-merge with a city-granular row from news.

The most valuable ten lines here are :func:`assert_headers`. Without them, an
ISO renaming a column produces a run that reads 40,000 rows, matches zero, and
exits 0 — the failure mode where you believe you have data and do not.
"""

from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
from collections.abc import Iterator, Mapping
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from tracker.ingest.iso_maps import MAPS_VERSION, VERIFIED_ISOS, IsoMap, get_map
from tracker.ingest.records import IngestRecord, IngestReport, SourceRecord
from tracker.models import utcnow
from tracker.normalize import (
    NormalizationError,
    is_blank,
    norm_date,
    norm_mw_detail,
    norm_phase,
    norm_state,
    norm_text,
)
from tracker.upsert import upsert_record
from tracker.vocab import DEFAULT_PHASE

log = logging.getLogger(__name__)

#: Rows per transaction. The PRD asks for chunks of 1000 for a 50 MB+ file.
CHUNK = 1000

#: Fail the run if more than this fraction of matched rows cannot be normalized.
#: A few bad rows are normal; a fifth of them means the mapping is wrong.
MAX_REJECT_RATE = 0.05

#: Phrases that actually name the facility type.
DC_PHRASES = (
    r"data\s*cent(?:er|re)",
    r"\bdatacent(?:er|re)\b",
    r"\bhyperscale\b",
    r"\bcolo(?:cation)?\b",
    r"\bcompute\s+campus\b",
)

#: Known data center operators and tenants, mapped to how their name should be
#: displayed. A weaker signal than a phrase match: "Microsoft Solar I LLC" is a
#: Microsoft-procured *generator*, not a data center.
#:
#: The display form matters because `company` is the primary column an operator
#: reads. Matching has to be lowercase, but writing the lowercase form back into
#: the database would render "coreweave" and "qts" in every table and export.
DC_OPERATOR_NAMES: dict[str, str] = {
    "microsoft": "Microsoft",
    "meta platforms": "Meta Platforms",
    "amazon": "Amazon",
    "aws": "Amazon",
    "google": "Google",
    "alphabet": "Google",
    "xai": "xAI",
    "x.ai": "xAI",
    "openai": "OpenAI",
    "stargate": "Stargate",
    "oracle": "Oracle",
    "coreweave": "CoreWeave",
    "qts": "QTS",
    "vantage": "Vantage Data Centers",
    "aligned": "Aligned Data Centers",
    "crusoe": "Crusoe",
    "digital realty": "Digital Realty",
    "equinix": "Equinix",
    "stack infra": "STACK Infrastructure",
    "switch inc": "Switch",
    "applied digital": "Applied Digital",
    "novva": "Novva",
    "cyrusone": "CyrusOne",
    "cloudhq": "CloudHQ",
    "edgeconnex": "EdgeConneX",
}

DC_OPERATORS = tuple(DC_OPERATOR_NAMES)


class HeaderError(ValueError):
    """The export does not have the columns this mapping expects."""


class IsoIngestError(ValueError):
    """The run cannot proceed or produced a result the operator must see."""


def _first(row: Mapping[str, Any], iso_map: IsoMap, logical: str) -> Any:
    """First non-blank value among a logical field's candidate columns."""
    spec = iso_map.columns.get(logical)
    if spec is None:
        return None
    for name in spec.src:
        if name in row and not is_blank(row[name]):
            return row[name]
    return None


# --- Readers ----------------------------------------------------------------


def _iter_csv(path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    """Stream a CSV. `utf-8-sig` first: ISO exports routinely carry a BOM."""
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
        handle.read(4096)
        handle.seek(0)
    except UnicodeDecodeError:
        log.warning("%s is not UTF-8; falling back to cp1252", path.name)
        handle = path.open("r", encoding="cp1252", newline="")
    with handle:
        # start=2 so the reported line number matches what an operator sees
        # in a spreadsheet, where row 1 is the header.
        yield from enumerate(csv.DictReader(handle), start=2)


def _iter_xlsx(path: Path, sheet: str | None) -> Iterator[tuple[int, dict[str, Any]]]:
    """Stream an xlsx. Two of the four ISOs publish xls/xlsx, not CSV."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised by the extras story
        raise IsoIngestError(
            f"{path.name} is a spreadsheet, which needs the `iso` extra:\n"
            '  python -m pip install -e ".[iso]"\n'
            "Alternatively, open it and Save As CSV."
        ) from exc

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = book[sheet] if sheet and sheet in book.sheetnames else book[book.sheetnames[0]]
        if sheet and sheet not in book.sheetnames:
            log.warning(
                "sheet %r not found in %s; using %r. Sheets: %s",
                sheet,
                path.name,
                ws.title,
                ", ".join(book.sheetnames),
            )
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows, ())]
        for lineno, values in enumerate(rows, start=2):
            if all(v is None for v in values):
                continue
            yield lineno, dict(zip(header, values, strict=False))
    finally:
        book.close()


def _iter_json(path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    """Stream a JSON payload. MISO publishes an API response, not a file export."""
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        for key in ("projects", "data", "Projects", "results"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
    if not isinstance(data, list):
        raise IsoIngestError(
            f"{path.name}: expected a JSON list of projects, or an object with a `projects` list"
        )
    for lineno, row in enumerate(data, start=1):
        if isinstance(row, dict):
            yield lineno, row


def iter_rows(path: Path, iso_map: IsoMap) -> Iterator[tuple[int, dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        yield from _iter_xlsx(path, iso_map.sheet)
    elif suffix == ".json":
        yield from _iter_json(path)
    else:
        yield from _iter_csv(path)


def assert_headers(path: Path, iso_map: IsoMap) -> list[str]:
    """Abort before row 1 if a required column is absent.

    This is the difference between "the ISO renamed a column and we ingested
    nothing" being a loud failure and being a silent success.
    """
    try:
        _, first_row = next(iter_rows(path, iso_map))
    except StopIteration:
        raise HeaderError(f"{path.name} has no data rows") from None

    present = set(first_row)
    missing = [
        logical
        for logical, spec in iso_map.columns.items()
        if spec.required and not (set(spec.src) & present)
    ]
    if missing:
        expected = {m: list(iso_map.columns[m].src) for m in missing}
        raise HeaderError(
            f"{path.name} is missing required column(s) for iso={iso_map.iso}.\n"
            f"  needed: {expected}\n"
            f"  found:  {sorted(present)}\n\n"
            "If the ISO renamed a column, map it without a code change:\n"
            '  --map-override \'{"<field>": ["<New Column Name>"]}\''
        )
    return sorted(present)


# --- Filtering --------------------------------------------------------------


def match_data_center(row: Mapping[str, Any], iso_map: IsoMap, mode: str) -> tuple[bool, str, int]:
    """Decide whether a queue row plausibly concerns a data center.

    Returns ``(keep, reason, confidence_cap)``.

    Modes:
        ``heuristic``  keyword match over name/entity columns. Cap 1.
        ``column:NAME=REGEX``  a real load-type column says so. Cap 2.
        ``none``  keep everything. Cap 0, for use with ``--dry-run``.
    """
    if mode == "none":
        return True, "unfiltered", 1

    if mode.startswith("column:"):
        spec = mode.removeprefix("column:")
        if "=" not in spec:
            raise IsoIngestError(
                'column filter must look like --filter "column:Load Type=(?i)data"'
            )
        column, pattern = spec.split("=", 1)
        value = str(row.get(column) or "")
        if column not in row:
            raise IsoIngestError(
                f"column filter names {column!r}, which is not in this file. Columns: {sorted(row)}"
            )
        # An authoritative load-type column is a real statement about the load,
        # not a guess, so it earns a higher cap.
        return bool(re.search(pattern, value, re.I)), f"column:{column}", 2

    haystack = " ".join(str(row.get(c) or "") for c in iso_map.dc_search_cols).lower()
    if any(re.search(p, haystack) for p in DC_PHRASES):
        return True, "name contains a data-center phrase", 1
    if any(op in haystack for op in DC_OPERATORS):
        return True, "name contains a known data-center operator", 1
    return False, "no-match", 0


# --- Row -> record ----------------------------------------------------------


def _infer_company(text: str, fallback: str) -> tuple[str, str | None]:
    """Recover the operator name from a queue row's free text.

    Returns ``(company, note)``.

    Why this exists: PJM's `Commercial Name` is usually a single-purpose entity
    ("Nova Solar I LLC") or a site label ("MS Mt Pleasant"), so using it as
    `company` splits one site across several projects — the PRD's own High risk
    example is exactly this, where "MS Mt Pleasant" and "Microsoft Racine County"
    must resolve to one Microsoft project. We already scan these columns for
    operator keywords in order to decide whether the row is a data center at all,
    so promoting the matched operator to `company` costs nothing and makes dedup
    work. The longest match wins so "meta platforms" beats a stray "meta".

    It is a heuristic, and the note records that it was applied.
    """
    haystack = text.lower()
    hits = [op for op in DC_OPERATORS if op in haystack]
    if not hits:
        return fallback, None
    best = max(hits, key=len)
    display = DC_OPERATOR_NAMES[best]
    if display.lower() == (fallback or "").lower():
        # The queue already named the operator correctly; disclosing an
        # "inference" that changed nothing is pure noise in the review output.
        return display, None
    return display, (
        f"company inferred as {display!r} from a keyword match on the queue row; "
        f"the queue's own commercial name was {fallback!r}"
    )


def _phase(row: Mapping[str, Any], iso_map: IsoMap) -> tuple[str | None, list[str]]:
    """Derive phase from queue status and the in-service/withdrawal dates.

    Returns ``(phase, notes)``. A ``None`` phase means the queue told us nothing,
    which the caller must translate into *omitting* `phase` from the claims so it
    is not presented as a cited fact.
    """
    notes: list[str] = []
    if not is_blank(_first(row, iso_map, "withdrawn")):
        return "cancelled", notes
    if not is_blank(_first(row, iso_map, "actual_online")):
        return "operational", notes

    raw = _first(row, iso_map, "raw_status")
    if is_blank(raw):
        return None, notes
    key = str(raw).strip().lower()
    if key in iso_map.status_map:
        return iso_map.status_map[key], notes
    try:
        resolved = norm_phase(raw)
    except NormalizationError:
        notes.append(f"queue status {str(raw)!r} did not map to a known phase")
        return None, notes
    return resolved, notes


def to_record(
    row: Mapping[str, Any],
    iso_map: IsoMap,
    *,
    fetched_at: datetime,
    reason: str,
    confidence_cap: int,
    trust_gen_mw: bool,
    file_digest: str,
    lineno: int,
) -> IngestRecord:
    """Build one IngestRecord from a queue row. Raises NormalizationError."""
    ext_id = norm_text(_first(row, iso_map, "ext_id")) or f"row{lineno}"
    state = norm_state(_first(row, iso_map, "state"), field="state")
    if not state:
        raise NormalizationError("state", _first(row, iso_map, "state"), "missing")

    county = norm_text(_first(row, iso_map, "county"))
    name = norm_text(_first(row, iso_map, "name")) or ext_id
    raw_company = norm_text(_first(row, iso_map, "company_raw")) or name

    notes: list[str] = []
    searchable = " ".join(
        str(row.get(c) or "") for c in (*iso_map.dc_search_cols, "Commercial Name")
    )
    company, company_note = _infer_company(searchable, raw_company)
    if company_note:
        notes.append(company_note)

    phase, phase_notes = _phase(row, iso_map)
    notes.extend(phase_notes)

    claims: dict[str, Any] = {
        "name": name,
        "company": company,
        "county": county,
        "state": state,
        "country": "US",
    }
    if phase is not None:
        claims["phase"] = phase

    # Generator capacity. Written to the project only on explicit opt-in.
    gen_mw = None
    raw_mw = _first(row, iso_map, "gen_mw")
    if not is_blank(raw_mw):
        parsed = norm_mw_detail(raw_mw, field="mw_planned")
        gen_mw = parsed.value
        if parsed.note:
            notes.append(parsed.note)

    fuel = norm_text(_first(row, iso_map, "fuel"))
    disclosure = (
        f"{iso_map.iso.upper()} queue {ext_id}: generation interconnection request "
        f"matched by {reason}"
    )
    if fuel:
        disclosure += f"; fuel={fuel}"
    if gen_mw is not None:
        if trust_gen_mw:
            claims["mw_planned"] = gen_mw
            disclosure += (
                f"; mw_planned={gen_mw:g} taken from generator nameplate via "
                "--trust-gen-mw, which is NOT confirmed data-center load"
            )
        else:
            disclosure += (
                f"; gen_queue_mw={gen_mw:g} (generator nameplate, not data-center "
                "load, so mw_planned left unset)"
            )
    notes.append(disclosure)

    for logical, column in (
        ("first_announced", "first_announced"),
        ("expected_online", "expected_online"),
    ):
        raw = _first(row, iso_map, column)
        if is_blank(raw):
            continue
        value = norm_date(raw, field=logical)
        if value is not None:
            claims[logical] = value

    excerpt = _excerpt(row, iso_map, ext_id)
    source = SourceRecord(
        # A fragment per queue id keeps the URL row-unique (so the
        # (project_id, url) constraint does not collapse many rows into one)
        # while still resolving to a real page a human can check.
        url=f"{iso_map.provenance_url}#{ext_id}",
        source_type=iso_map.source_type,
        fetched_at=fetched_at,
        excerpt=excerpt,
        claims=claims,
        extractor=f"{iso_map.iso}:{MAPS_VERSION}:sha256={file_digest}:row={lineno}",
    )

    project = {
        "name": name,
        "company": company,
        "county": county,
        "state": state,
        "country": "US",
        "phase": phase or DEFAULT_PHASE,
    }
    return IngestRecord(
        project=project,
        sources=[source],
        notes=notes,
        confidence_cap=confidence_cap,
    )


def _excerpt(row: Mapping[str, Any], iso_map: IsoMap, ext_id: str) -> str:
    """A deterministic rendering of the load-bearing cells.

    A CSV row has no prose to quote, so the citation's "excerpt" is the row
    itself, rendered stably. That keeps the traceability promise honest: the
    excerpt is exactly what the file said.
    """
    parts = [f"{iso_map.iso.upper()} queue {ext_id}"]
    for label, logical in (
        ("name", "name"),
        ("entity", "company_raw"),
        ("county", "county"),
        ("state", "state"),
        ("MW", "gen_mw"),
        ("fuel", "fuel"),
        ("status", "raw_status"),
        ("submitted", "first_announced"),
    ):
        value = _first(row, iso_map, logical)
        if not is_blank(value):
            parts.append(f"{label}={value}")
    return " | ".join(parts)[:500]


# --- Run --------------------------------------------------------------------


def _digest(path: Path) -> str:
    sha = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1 << 20), b""):
            sha.update(block)
    return sha.hexdigest()[:12]


def run(
    session: Session,
    path: Path,
    *,
    iso: str = "pjm",
    filter_mode: str = "heuristic",
    trust_gen_mw: bool = False,
    dry_run: bool = False,
    limit: int | None = None,
    rejects_out: Path | None = None,
    map_override: dict[str, list[str]] | None = None,
    force_new: bool = False,
) -> IngestReport:
    """Load an ISO queue export.

    A row that fails normalization is logged and counted; it never aborts the
    run. The run *does* fail (via :class:`IsoIngestError`) when zero rows matched
    the filter or when the reject rate exceeds :data:`MAX_REJECT_RATE`, because
    both mean the output is not what the operator thinks it is.
    """
    iso_map = get_map(iso, map_override)
    if iso_map.iso not in VERIFIED_ISOS:
        log.warning(
            "column names for iso=%s are unverified assumptions; if this run "
            "aborts on missing headers, use --map-override to correct them",
            iso_map.iso,
        )

    assert_headers(path, iso_map)
    file_digest = _digest(path)
    fetched_at = utcnow()
    report = IngestReport()
    batch: list[IngestRecord] = []
    reject_reasons: list[str] = []

    def flush() -> None:
        for record in batch:
            result = upsert_record(session, record, force_new=force_new)
            report.bump(result.action)
            report.events += result.events_written
            report.conflicts += len(result.conflicts)
            if result.duplicate_of is not None:
                report.duplicates_flagged += 1
        if dry_run:
            session.rollback()
        else:
            session.commit()
        batch.clear()

    for lineno, row in iter_rows(path, iso_map):
        # Checked before filtering: --limit means "read at most N rows", and a
        # filtered row still consumed one.
        if limit is not None and report.read >= limit:
            break
        report.read += 1
        keep, reason, cap = match_data_center(row, iso_map, filter_mode)
        if not keep:
            report.filtered += 1
            continue
        try:
            record = to_record(
                row,
                iso_map,
                fetched_at=fetched_at,
                reason=reason,
                confidence_cap=cap,
                trust_gen_mw=trust_gen_mw,
                file_digest=file_digest,
                lineno=lineno,
            )
        except NormalizationError as exc:
            report.rejected += 1
            reject_reasons.append(exc.reason)
            log.warning(
                "REJECT iso=%s line=%d id=%r field=%s value=%r reason=%s",
                iso_map.iso,
                lineno,
                _first(row, iso_map, "ext_id"),
                exc.field,
                exc.value,
                exc.reason,
            )
            if rejects_out:
                _append_reject(rejects_out, lineno, row, exc)
            continue

        batch.append(record)
        if len(batch) >= CHUNK:
            flush()

    flush()
    _check_run_quality(report, iso_map, filter_mode, reject_reasons)
    return report


def _check_run_quality(
    report: IngestReport, iso_map: IsoMap, filter_mode: str, reject_reasons: list[str]
) -> None:
    """Turn two silent-success failure modes into loud ones."""
    matched = report.read - report.filtered
    if report.read and matched == 0:
        raise IsoIngestError(
            f"read {report.read} row(s) from the {iso_map.iso.upper()} queue and "
            f"matched none with filter={filter_mode!r}.\n"
            "That usually means the columns searched for data-center keywords are "
            f"absent or renamed (searched: {list(iso_map.dc_search_cols)}).\n"
            "Check with --filter none --dry-run --limit 5 to see real rows."
        )
    if matched and report.rejected / matched > MAX_REJECT_RATE:
        top = ", ".join(sorted(set(reject_reasons))[:3])
        raise IsoIngestError(
            f"{report.rejected} of {matched} matched row(s) failed normalization "
            f"({report.rejected / matched:.0%}, limit {MAX_REJECT_RATE:.0%}).\n"
            f"Most common reasons: {top}\n"
            "The column mapping is probably wrong for this file."
        )


def _append_reject(out: Path, lineno: int, row: Mapping[str, Any], exc: NormalizationError) -> None:
    """Append a rejected row as JSONL so it can be fixed and re-run."""
    out.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "line": lineno,
        "field": exc.field,
        "value": str(exc.value),
        "reason": exc.reason,
        "row": {k: (None if v is None else str(v)) for k, v in row.items()},
    }
    with out.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")


__all__ = [
    "CHUNK",
    "DC_OPERATORS",
    "DC_PHRASES",
    "MAX_REJECT_RATE",
    "HeaderError",
    "IsoIngestError",
    "assert_headers",
    "iter_rows",
    "match_data_center",
    "run",
    "to_record",
]
